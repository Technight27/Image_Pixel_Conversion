#!/usr/bin/env python
# coding: utf-8

# In[1]:


from PIL import Image
import openpyxl

try:
    # Load image
    image_path = 'C:\\Users\\Sayan\\Downloads\\sample.jpg'
    original_image = Image.open(image_path)

    # Convert image to grayscale for simplicity
    gray_image = original_image.convert('L')

    # Get pixel values as a 2D matrix
    pixel_matrix = list(gray_image.getdata())
    width, height = gray_image.size
    pixel_matrix = [pixel_matrix[i:i + width] for i in range(0, len(pixel_matrix), width)]

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write pixel values to Excel sheet
    for y in range(height):
        for x in range(width):
            pixel_value = pixel_matrix[y][x]
            ws.cell(row=y+1, column=x+1, value=pixel_value)

    # Save the Excel workbook
    excel_file = 'C:\\Users\\Sayan\\Downloads\\pixel_values.xlsx'
    wb.save(excel_file)
    print(f"Pixel values saved to {excel_file}")

except Exception as e:
    print(f"Error: {e}")


# In[2]:


from PIL import Image
import openpyxl

try:
    # Load Excel workbook
    excel_file = 'C:\\Users\\Sayan\\Downloads\\pixel_values.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Get pixel values from Excel sheet
    pixel_matrix = []
    for row in ws.iter_rows(values_only=True):
        pixel_matrix.append(list(row))

    # Reconstruct the image from the matrix
    width = len(pixel_matrix[0])
    height = len(pixel_matrix)
    reconstructed_image = Image.new('L', (width, height))
    reconstructed_image.putdata([pixel for row in pixel_matrix for pixel in row])

    # Save or display the reconstructed image
    reconstructed_image.save('reconstructed_image.jpg')
    reconstructed_image.show()  # Display the reconstructed image
    print("Image reconstruction successful.")

except Exception as e:
    print(f"Error: {e}")


# In[ ]:




